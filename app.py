from flask import Flask, render_template, request, jsonify, send_file, session
import joblib
import numpy as np
import os
import json
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'sapphire-residence-tegal-2026'

# -------------------------------------------------------------------
# KONFIGURASI
# -------------------------------------------------------------------
BASE_PATH = os.path.dirname(os.path.abspath(__file__))
MODEL_PATH = os.path.join(BASE_PATH, 'models')
DATASET_PATH = os.path.join(BASE_PATH, 'Dataset_Sapphire_Final.csv')

# -------------------------------------------------------------------
# INISIALISASI GLOBAL
# -------------------------------------------------------------------
scaler = None
le_blok = None
le_posisi = None
models = {}
training_results = {}
resources_loaded = False


def load_resources():
    """Muat semua model, encoder, dan hasil training dari disk."""
    global scaler, le_blok, le_posisi, models, training_results, resources_loaded

    try:
        print(f"Memuat resources dari: {MODEL_PATH}")

        scaler = joblib.load(os.path.join(MODEL_PATH, 'scaler.pkl'))
        le_blok = joblib.load(os.path.join(MODEL_PATH, 'le_blok.pkl'))
        le_posisi = joblib.load(os.path.join(MODEL_PATH, 'le_posisi.pkl'))

        models = {
            'adam': joblib.load(os.path.join(MODEL_PATH, 'model_adam.pkl')),
            'sgd': joblib.load(os.path.join(MODEL_PATH, 'model_sgd.pkl')),
            'lbfgs': joblib.load(os.path.join(MODEL_PATH, 'model_lbfgs.pkl'))
        }

        # Muat hasil training (metrik, statistik, deskriptif)
        json_path = os.path.join(MODEL_PATH, 'training_results.json')
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                training_results = json.load(f)
            print("Hasil training berhasil dimuat dari JSON.")
        else:
            print("WARNING: training_results.json tidak ditemukan.")

        resources_loaded = True
        print("Semua resources berhasil dimuat.")
        return True

    except Exception as e:
        print(f"ERROR saat memuat resources: {e}")
        resources_loaded = False
        return False


load_resources()


# -------------------------------------------------------------------
# FUNGSI UTILITAS
# -------------------------------------------------------------------
def hitung_cicilan(plafon, bunga_tahunan, tenor_tahun):
    """Hitung cicilan bulanan menggunakan formula anuitas."""
    r = bunga_tahunan / 12
    n = tenor_tahun * 12
    if r == 0:
        return plafon / n
    return plafon * (r * (1 + r)**n) / ((1 + r)**n - 1)


def format_rupiah(angka):
    """Format angka ke format Rupiah Indonesia."""
    return f"{angka:,.0f}".replace(",", ".")


def analisis_dsr(gaji, cicilan):
    """Analisis Debt Service Ratio (DSR) 30%."""
    if gaji <= 0:
        return {
            'rasio': 0,
            'status': 'DATA TIDAK LENGKAP',
            'kelas': 'warning',
            'saran': 'Masukkan data penghasilan untuk analisis kelayakan kredit.'
        }

    rasio = (cicilan / gaji) * 100

    if rasio <= 30:
        return {
            'rasio': round(rasio, 2),
            'status': 'LAYAK',
            'kelas': 'success',
            'saran': f'Rasio cicilan terhadap penghasilan Anda sebesar {rasio:.1f}%, masih di bawah ambang batas DSR 30%. Kredit dinilai layak secara finansial.'
        }
    elif rasio <= 40:
        return {
            'rasio': round(rasio, 2),
            'status': 'PERTIMBANGKAN KEMBALI',
            'kelas': 'warning',
            'saran': f'Rasio cicilan terhadap penghasilan Anda sebesar {rasio:.1f}%, melebihi ambang batas ideal DSR 30%. Pertimbangkan untuk memilih tenor lebih panjang atau unit dengan harga lebih rendah.'
        }
    else:
        return {
            'rasio': round(rasio, 2),
            'status': 'TIDAK LAYAK',
            'kelas': 'danger',
            'saran': f'Rasio cicilan terhadap penghasilan Anda sebesar {rasio:.1f}%, jauh melebihi batas aman DSR 30%. Disarankan untuk memilih unit yang lebih terjangkau atau menambah uang muka.'
        }


# -------------------------------------------------------------------
# ROUTES
# -------------------------------------------------------------------

@app.route('/')
def index():
    """Halaman utama - Form prediksi harga."""
    daftar_blok = sorted(le_blok.classes_.tolist()) if le_blok else []
    return render_template('index.html',
                           page='prediksi',
                           daftar_blok=daftar_blok)


@app.route('/predict', methods=['POST'])
def predict():
    """Proses prediksi harga dan analisis kelayakan kredit."""
    if not resources_loaded:
        return render_template('index.html',
                               page='prediksi',
                               error="Model belum dimuat. Jalankan train_model.py terlebih dahulu.")

    try:
        # 1. Ambil input dari form
        nama_blok = request.form['nama_blok']
        lb = float(request.form['luas_bangunan'])
        lt = float(request.form['luas_tanah'])
        kt = int(request.form['kamar_tidur'])
        km = int(request.form['kamar_mandi'])
        posisi = request.form['posisi']
        sdl = int(request.form['smart_lock'])
        opt = request.form['optimizer']
        gaji = float(request.form.get('gaji', 0))

        # 2. Transform input
        blok_encoded = le_blok.transform([nama_blok])[0]
        posisi_encoded = le_posisi.transform([posisi])[0]
        input_raw = np.array([[blok_encoded, lb, lt, kt, km, posisi_encoded, sdl]])
        input_scaled = scaler.transform(input_raw)

        # 3. Prediksi (kembalikan dari skala jutaan ke Rupiah)
        prediksi_mentah = models[opt].predict(input_scaled)[0]
        prediksi = max(0, prediksi_mentah * 1000000)

        # 4. Ambil metrik model yang digunakan
        metrik = training_results.get('metrik_model', {}).get(opt.upper(), {})

        # 5. Simulasi KPR
        dp_persen = 10  # DP 10%
        dp = prediksi * (dp_persen / 100)
        plafon = prediksi - dp
        bunga = 0.08  # 8% per tahun

        cicilan_10 = hitung_cicilan(plafon, bunga, 10)
        cicilan_15 = hitung_cicilan(plafon, bunga, 15)
        cicilan_20 = hitung_cicilan(plafon, bunga, 20)

        # 6. Analisis DSR untuk masing-masing tenor
        dsr_10 = analisis_dsr(gaji, cicilan_10)
        dsr_15 = analisis_dsr(gaji, cicilan_15)
        dsr_20 = analisis_dsr(gaji, cicilan_20)

        # Gunakan tenor 20 tahun sebagai acuan utama
        status_utama = dsr_20

        # 7. Render hasil
        daftar_blok = sorted(le_blok.classes_.tolist())

        return render_template('index.html',
                               page='prediksi',
                               daftar_blok=daftar_blok,
                               hasil_harga=format_rupiah(prediksi),
                               harga_raw=prediksi,
                               optimizer_nama=opt.upper(),
                               metrik=metrik,
                               dp=format_rupiah(dp),
                               dp_persen=dp_persen,
                               plafon=format_rupiah(plafon),
                               bunga=bunga * 100,
                               cicilan_10=format_rupiah(cicilan_10),
                               cicilan_15=format_rupiah(cicilan_15),
                               cicilan_20=format_rupiah(cicilan_20),
                               dsr_10=dsr_10,
                               dsr_15=dsr_15,
                               dsr_20=dsr_20,
                               status_utama=status_utama,
                               input_user=request.form)

    except Exception as e:
        daftar_blok = sorted(le_blok.classes_.tolist()) if le_blok else []
        return render_template('index.html',
                               page='prediksi',
                               daftar_blok=daftar_blok,
                               error=str(e))


@app.route('/dashboard')
def dashboard():
    """Halaman perbandingan performa optimizer."""
    metrik = training_results.get('metrik_model', {})
    arsitektur = training_results.get('arsitektur', {})
    best = training_results.get('optimizer_terbaik', 'LBFGS')
    split = training_results.get('split_data', {})
    return render_template('dashboard.html',
                           page='dashboard',
                           metrik=metrik,
                           arsitektur=arsitektur,
                           optimizer_terbaik=best,
                           split_data=split)


@app.route('/analisis')
def analisis():
    """Halaman analisis deskriptif dan uji statistik."""
    deskriptif = training_results.get('deskriptif', {})
    statistik = training_results.get('statistik', {})
    arsitektur = training_results.get('arsitektur', {})
    fitur = training_results.get('fitur', [])
    return render_template('analisis.html',
                           page='analisis',
                           deskriptif=deskriptif,
                           statistik=statistik,
                           arsitektur=arsitektur,
                           fitur=fitur)


@app.route('/api/metrics')
def api_metrics():
    """API endpoint untuk data metrik (JSON)."""
    return jsonify(training_results)


# -------------------------------------------------------------------
# UTILITAS EXCEL
# -------------------------------------------------------------------
def style_excel(ws, title, headers, data_rows, start_row=1):
    """Styling Excel worksheet secara konsisten."""
    hdr_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    hdr_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Calibri', size=10)
    cell_align = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Title
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(name='Calibri', bold=True, size=13, color='1E3A5F')
    title_cell.alignment = Alignment(horizontal='left')

    # Headers
    hr = start_row + 1
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin_border

    # Data
    for ri, row_data in enumerate(data_rows, hr + 1):
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = cell_font
            c.alignment = cell_align
            c.border = thin_border
            if ri % 2 == 0:
                c.fill = PatternFill(start_color='F5F6F8', end_color='F5F6F8', fill_type='solid')

    # Auto-width
    for ci in range(1, len(headers) + 1):
        max_len = len(str(headers[ci - 1]))
        for ri in range(hr + 1, hr + 1 + len(data_rows)):
            val = ws.cell(row=ri, column=ci).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 35)

    return hr + 1 + len(data_rows)


# -------------------------------------------------------------------
# EXPORT ROUTES
# -------------------------------------------------------------------
@app.route('/export/prediksi', methods=['POST'])
def export_prediksi():
    """Export hasil prediksi ke Excel."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Hasil Prediksi'

        # Data dari form
        blok = request.form.get('blok', '-')
        lb = request.form.get('lb', '-')
        lt = request.form.get('lt', '-')
        kt = request.form.get('kt', '-')
        km = request.form.get('km', '-')
        posisi = request.form.get('posisi', '-')
        sdl = request.form.get('sdl', '-')
        optimizer = request.form.get('optimizer', '-')
        harga = request.form.get('harga', '-')
        dp = request.form.get('dp', '-')
        plafon = request.form.get('plafon', '-')
        cicilan_10 = request.form.get('cicilan_10', '-')
        cicilan_15 = request.form.get('cicilan_15', '-')
        cicilan_20 = request.form.get('cicilan_20', '-')
        gaji = request.form.get('gaji', '-')
        dsr_status = request.form.get('dsr_status', '-')
        dsr_rasio = request.form.get('dsr_rasio', '-')

        # Spesifikasi unit
        next_row = style_excel(ws, 'SPESIFIKASI UNIT',
            ['Parameter', 'Nilai'],
            [
                ['Blok Kavling', blok],
                ['Luas Bangunan (m2)', lb],
                ['Luas Tanah (m2)', lt],
                ['Kamar Tidur', kt],
                ['Kamar Mandi', km],
                ['Posisi Kavling', posisi],
                ['Smart Door Lock', 'Ya' if sdl == '1' else 'Tidak'],
                ['Optimizer', optimizer],
            ])

        # Hasil prediksi
        next_row = style_excel(ws, 'HASIL PREDIKSI HARGA',
            ['Keterangan', 'Nilai'],
            [
                ['Estimasi Harga KPR', f'Rp {harga}'],
                ['Uang Muka (DP 10%)', f'Rp {dp}'],
                ['Plafon Kredit', f'Rp {plafon}'],
                ['Bunga', '8.0% / tahun'],
            ], start_row=next_row + 1)

        # Simulasi KPR
        next_row = style_excel(ws, 'SIMULASI KREDIT',
            ['Tenor', 'Cicilan per Bulan'],
            [
                ['10 Tahun', f'Rp {cicilan_10}'],
                ['15 Tahun', f'Rp {cicilan_15}'],
                ['20 Tahun', f'Rp {cicilan_20}'],
            ], start_row=next_row + 1)

        # Kelayakan
        next_row = style_excel(ws, 'ANALISIS KELAYAKAN KREDIT (DSR)',
            ['Parameter', 'Nilai'],
            [
                ['Penghasilan Bulanan', f'Rp {gaji}'],
                ['DSR (Tenor 20 Tahun)', f'{dsr_rasio}%'],
                ['Status Kelayakan', dsr_status],
                ['Ambang Batas DSR', '30%'],
            ], start_row=next_row + 1)

        # Timestamp
        r = next_row + 2
        ws.cell(row=r, column=1, value=f'Dicetak pada: {datetime.now().strftime("%d %B %Y, %H:%M WIB")}')
        ws.cell(row=r, column=1).font = Font(name='Calibri', size=9, italic=True, color='888888')

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        fname = f'Prediksi_Sapphire_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return send_file(output, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/export/dashboard')
def export_dashboard():
    """Export perbandingan optimizer ke Excel."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Perbandingan Optimizer'
        metrik = training_results.get('metrik_model', {})
        arsitektur = training_results.get('arsitektur', {})

        # Arsitektur info
        next_row = style_excel(ws, 'ARSITEKTUR MODEL ANN',
            ['Parameter', 'Nilai'],
            [
                ['Hidden Layers', '-'.join(str(x) for x in arsitektur.get('hidden_layers', []))],
                ['Fitur Input', str(arsitektur.get('input_layer', 7))],
                ['Fungsi Aktivasi', arsitektur.get('aktivasi', 'ReLU')],
                ['Max Iterasi', str(arsitektur.get('max_iter', 10000))],
                ['Data Training', str(training_results.get('split_data', {}).get('training', '-'))],
                ['Data Testing', str(training_results.get('split_data', {}).get('testing', '-'))],
                ['Rasio Split', training_results.get('split_data', {}).get('ratio', '80:20')],
            ])

        # Tabel metrik
        rows = []
        for opt_name, m in metrik.items():
            rows.append([
                opt_name,
                round(m.get('r2', 0), 6),
                f"Rp {m.get('mae', 0):,.0f}".replace(',', '.'),
                f"Rp {m.get('rmse', 0):,.0f}".replace(',', '.'),
                f"{m.get('mape', 0):.4f}%",
                f"{m.get('akurasi', 0):.4f}%",
                m.get('n_iter', '-'),
            ])

        next_row = style_excel(ws, 'PERBANDINGAN METRIK EVALUASI',
            ['Optimizer', 'R2 Score', 'MAE', 'RMSE', 'MAPE', 'Akurasi', 'Iterasi'],
            rows, start_row=next_row + 1)

        # Kesimpulan
        best = training_results.get('optimizer_terbaik', '-')
        r = next_row + 1
        ws.cell(row=r, column=1, value=f'Optimizer Terbaik: {best}')
        ws.cell(row=r, column=1).font = Font(name='Calibri', bold=True, size=11, color='0D7C66')

        r += 2
        ws.cell(row=r, column=1, value=f'Dicetak pada: {datetime.now().strftime("%d %B %Y, %H:%M WIB")}')
        ws.cell(row=r, column=1).font = Font(name='Calibri', size=9, italic=True, color='888888')

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        fname = f'Dashboard_Sapphire_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return send_file(output, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/export/analisis')
def export_analisis():
    """Export analisis statistik ke Excel."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Analisis Statistik'
        deskriptif = training_results.get('deskriptif', {})
        statistik = training_results.get('statistik', {})

        # Statistik deskriptif
        desc_rows = []
        for nama, stats in deskriptif.get('statistik', {}).items():
            desc_rows.append([
                nama,
                f"{stats['mean']:,.2f}".replace(',', '.'),
                f"{stats['std']:,.2f}".replace(',', '.'),
                f"{stats['min']:,.0f}".replace(',', '.'),
                f"{stats['median']:,.0f}".replace(',', '.'),
                f"{stats['max']:,.0f}".replace(',', '.'),
            ])

        next_row = style_excel(ws, 'STATISTIK DESKRIPTIF',
            ['Variabel', 'Mean', 'Std. Dev', 'Min', 'Median', 'Max'],
            desc_rows)

        # Uji F
        uji_f = statistik.get('uji_f', {})
        next_row = style_excel(ws, 'UJI F (SIMULTAN)',
            ['Parameter', 'Nilai'],
            [
                ['F-Statistic', f"{uji_f.get('f_statistic', 0):,.4f}".replace(',', '.')],
                ['P-Value', f"{uji_f.get('p_value', 0):.2e}"],
                ['Kesimpulan', 'Signifikan' if uji_f.get('signifikan') else 'Tidak Signifikan'],
                ['R-Squared', f"{statistik.get('r_squared', 0):.6f}"],
                ['Adj. R-Squared', f"{statistik.get('adj_r_squared', 0):.6f}"],
            ], start_row=next_row + 1)

        # Uji t
        t_rows = []
        for item in statistik.get('uji_t', []):
            if item.get('keterangan'):
                t_rows.append([item['variabel'], '--', '--', '--', item['keterangan']])
            else:
                t_rows.append([
                    item['variabel'],
                    f"{item['koefisien']:,.4f}".replace(',', '.'),
                    f"{item['t_statistic']:.4f}",
                    f"{item['p_value']:.6f}",
                    'Signifikan' if item.get('signifikan') else 'Tidak Signifikan',
                ])

        next_row = style_excel(ws, 'UJI T (PARSIAL)',
            ['Variabel', 'Koefisien', 't-Statistic', 'P-Value', 'Status'],
            t_rows, start_row=next_row + 1)

        # Distribusi
        dist_rows = []
        for tipe, info in deskriptif.get('distribusi_tipe', {}).items():
            dist_rows.append([tipe, info['count'], f"{info['persen']}%"])

        next_row = style_excel(ws, 'DISTRIBUSI TIPE RUMAH',
            ['Tipe', 'Jumlah Unit', 'Persentase'],
            dist_rows, start_row=next_row + 1)

        r = next_row + 1
        ws.cell(row=r, column=1, value=f'Dicetak pada: {datetime.now().strftime("%d %B %Y, %H:%M WIB")}')
        ws.cell(row=r, column=1).font = Font(name='Calibri', size=9, italic=True, color='888888')

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        fname = f'Analisis_Sapphire_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return send_file(output, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# -------------------------------------------------------------------
# MAIN
# -------------------------------------------------------------------
if __name__ == '__main__':
    app.run(debug=True, port=5000)